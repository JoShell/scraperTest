import openpyxl, os # import os and excel modules

wb = openpyxl.Workbook() # create a new workbook and assign it to a variable - workbook only exists in RAM
wb

sheet = wb['Sheet'] # inside the workbook object there is the sheet object, called sheet, that we now assign the variable name sheet

sheet['A1'].value # to know the value of the cell A1 inside the sheet object - cells also have a string value

sheet['A1'] = 42 # assign a value to the A1 cell object in sheet

wb.save('example.xlsx') # save the workbook object to disk with a file name

# make sure to save to a new location each time

sheet2 = wb.create_sheet() # create a new sheet object inside the work book object
sheet2.title = 'My new sheet name' # assign a name to the new sheet object

wb.save('example.xlsx') # save the sheet object update

wb.create_sheet(index=0, title='Spam eggs and spam') # you can update sheets by location, 0 is you want to keep adding to the begining

wb.save('example.xlsx') # save the sheet object update